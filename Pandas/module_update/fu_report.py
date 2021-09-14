
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# mpath1 = r'D:\Study\15.Python\Pandas\excel_file\FU-WEC-C-1000-4534.xlsx'
# mpath2 = r'D:\Study\15.Python\Pandas\excel_file\FU-WEC-C-7000-4535.xlsx'
# mpath = [mpath1,mpath2]

# Tạo dataframe cho nội dung các báo cáo
def create_fu_data(path):
    
    df = pd.read_excel(path, sheet_name=0, usecols=range(1,10),
    skipfooter=5, skiprows=16, dtype='str')
    df = df.dropna(how = 'all')
    # df = df.fillna(0)
    df['joint_id'] = df['Spool No.']+'\\'+df['Iso. Dwg. No.']+'\\'+df['Joint No.']+'\\'+df['DB']
    df['DB'] = df['DB'].astype('float')
    
    df = df.reset_index(drop=True)
    return df

# Tạo dataframe cho ngày fit-up và số báo cáo
def list_date_report(path):
    wb= load_workbook(path)
    ws = wb.worksheets[0]
    date_report = ws['E15'].value
    return date_report

def list_report_no(path):
    wb= load_workbook(path)
    ws = wb.worksheets[0]
    report_no = ws['I7'].value
    return report_no
