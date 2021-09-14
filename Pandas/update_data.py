
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
from tkinter import filedialog
# Chèn module đã tạo vào trong scrip
from module_update import fu_report, fu_data

# tạo dataframe cho joint data
data_path = r'D:\Study\15.Python\Pandas\excel_file\CMS-Pipe-Test.xlsm'
data_df = fu_data.create_joint_data(data_path)

#  Đặt đường dẫn đến các file FU và file data
mFolder = filedialog.askdirectory(title='Chon Folder Fit-up update')
mpath = []
for links, folders, files in os.walk(mFolder):
    for file in files:
        f = [os.path.join(links,file)]
        mpath.extend(f)

# Tạo dataframe cho FU
fu_df = []
for path in mpath:
    df = fu_report.create_fu_data(path)
    fu_df.append(df)


# So sánh các DF
def compare_fun(report_dataFrame,data_dataFrame):
    joint_bc=[]
    joint_trung=[]
    joint_koupdate = []
    for i in report_dataFrame['joint_id'].index:
        data_df_ss= data_dataFrame[data_dataFrame['joint_id']==report_dataFrame['joint_id'][i]]
        # kiểm tra nếu data_df_ss là rỗng thì chuyển ngay vào list không update được
        if len(data_df_ss.index)==0:
            joint_koupdate.append(i)
        elif pd.isna(data_df_ss['Report_FU'][data_df_ss['joint_id'].index[0]]):
            joint_bc.append((data_df_ss['joint_id'].index[0]+7))
        else:
            joint_trung.append(data_df_ss['joint_id'].index[0])
    return joint_bc, joint_trung, joint_koupdate

#  So sánh các dữ liệu
list_compare = []
for df_rp in fu_df:
    t = compare_fun(df_rp,data_df)
    list_compare.append(t)

# Tạo list dữ liệu có thể update và dữ liệu trùng là các vị trí dòng trong file data
list_update = []
list_trung = []
list_ko_update=[]
for compare in list_compare:
    # list_update là list của list
    list_update.append(list(compare)[0])
    # List_trung là list của các dòng trong file data
    list_trung = list_trung + list(compare)[1]
    # list_ko_update là list của list
    list_ko_update.append(list(compare)[2])
  
#  Tạo list dataframe theo các báo cáo ko update được
if list_ko_update!=[]:
    list_df_ko_update = []
    for i in range(len(fu_df)):
        dfi = fu_df[i].filter(items=list_ko_update[i], axis= 0)
        list_df_ko_update.append(dfi)
    #  Nối các dataframe không update được lại với nhau và in thành file
    df_ko_update = pd.concat(list_df_ko_update, axis=0)
    df_ko_update.to_excel('ko_update_duoc.xlsx')

# Tạo ra dataframe dữ liệu trùng và in thành file
if len(list_trung)!=0:
    df_trung = data_df.filter(items=list_trung, axis = 0)
    df_trung.to_excel('trung_joint.xlsx')

# Dùng openpyxl để  điền dữ liệu vào trong file data
# Lấy dữ liệu ngày tháng số report
date_rp=[]
num_rp = []
for rp_path in mpath:
    wb = load_workbook(rp_path)
    ws = wb.worksheets[0]
    date_rp.append(ws['E15'].value)
    num_rp.append(ws['I7'].value)

# Load workbook data  và điền dữ liệu update vào file (nhớ phải lưu vào)
wb = load_workbook(data_path, keep_vba=True, read_only=False)
ws = wb.worksheets[0]
for up_date in range(len(list_update)):
    for u in list_update[up_date]:
        ws.cell(row=u, column=21, value = date_rp[up_date])
        ws.cell(row=u, column=22, value = num_rp[up_date])
wb.save(data_path)
