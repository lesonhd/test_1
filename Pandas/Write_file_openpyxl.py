import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date

my_data = r'D:\Study\15.Python\Pandas\test.xlsm'

m_report = r'D:\Study\15.Python\Pandas\excel_file\FU-WEC-C-7000-4535.xlsx'

m_list =[2,4,5,7]

# Lấy ngày báo cáo và số báo cáo
wb_rp = load_workbook(m_report)
ws_rp = wb_rp.worksheets[0]
rp_date = ws_rp['E15'].value
rp_date = date.isoformat(rp_date)   # định dạng lại ngày tháng cho đẹp
rp_num = ws_rp['I7'].value

# Mở workbook data và điền dữ liệu vào
wb_data = load_workbook(my_data, keep_vba= True, read_only = False)
ws_data = wb_data.worksheets[0]
for m_row in m_list:
    ws_data.cell(row=m_row, column=5, value=rp_date)
    ws_data.cell(row=m_row, column=6, value=rp_num)

# Lưu lại file vừa điền
wb_data.save(my_data)
