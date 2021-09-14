import pandas as pd
import numpy as np
import openpyxl 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Lấy dữ liệu từ CMS
mPath = r'E:\\Onedrive\\9.Packet A1\\9.QA-QC-WE\\WEC-CMS-Flare.xlsx'
Cot_sudung = [0,1,2,10,5,12,8,19,14,17,18,21]
df = pd.read_excel(mPath,sheet_name=0, skiprows=5, usecols=Cot_sudung)

#  Tạo ra các cột mới
df['Drawing_sheet'] = df['Drawing No']+'-Sheet-'+df['Sheet No']
df['Trace-before'] = ''
df['Trace-After'] = ''
df['Spool'] = ''
df['Fit-up-A'] = 'A'
df['Welding-A'] = 'A'
df['Gap'] = 'N/A'
df['Signed-A'] = 'A'

# Sắp xếp lại các cột và loại bỏ cột không cần thiết
df1 = df[df.columns[[12,2,15,5,3,6,13,14,4,16,17,10,18,7,8,9,19,11]]]

# Lấy danh sách các số báo cáo cần tạo
Path_report_list = r'E:\\Onedrive\\9.Packet A1\\9.QA-QC-WE\\2.Visual\\List_report_python.xlsm'
df_report = pd.read_excel(Path_report_list)
report_dic = df_report.to_dict()
report_list = list(report_dic['list_report'].values())

def tao_bao_cao(report_1):
    # Làm thử 1 báo cáo
    # report_1 = report_list[0]
    df_rp1 = df1.loc[df1['Visual Report']==report_1]
    rows = dataframe_to_rows(df_rp1,index=False, header=False)
    # Load from vào openpyxl
    m_path_form = r'E:\\Onedrive\\9.Packet A1\\9.QA-QC-WE\\2.Visual\\Welding Form W24.xlsx'
    wb = load_workbook(m_path_form, read_only= False)
    ws = wb.worksheets[0]
    # Xóa dòng thừa
    ws.delete_rows(208-(200-len(df_rp1.index)),200-len(df_rp1.index)-1)
    i = 9
    for row in rows:
        for y in range(1,18):
            ws.cell(column=y,row=i, value=row[y-1])
        i = i+1

    ws.cell(column=17, row= 4, value= row[17])
    ws.cell(column=2, row= 5, value= row[11])
    # chỉnh lại chỗ note
    ws.merge_cells(start_row=(208-(200-len(df_rp1.index))+2),start_column= 1 , 
        end_row=(208-(200-len(df_rp1.index))+2), end_column=18)
    ws.row_dimensions[208-(200-len(df_rp1.index))+2].height = 70

    # Chỉnh lại chỗ chữ ký
    for i in range(3,7):
        ws.merge_cells(start_row=(208-(200-len(df_rp1.index))+i),start_column= 2 , 
        end_row=(208-(200-len(df_rp1.index))+i), end_column=7)

        ws.merge_cells(start_row=(208-(200-len(df_rp1.index))+i),start_column= 8 , 
        end_row=(208-(200-len(df_rp1.index))+i), end_column=13)

        ws.merge_cells(start_row=(208-(200-len(df_rp1.index))+i),start_column= 14 , 
        end_row=(208-(200-len(df_rp1.index))+i), end_column=18)

    wb.save(r'E:\\Onedrive\\9.Packet A1\\9.QA-QC-WE\\2.Visual' + '\\\\'+ report_1[12:] +'.xlsx')
    wb.close()

for rp in report_list:
    tao_bao_cao(rp)

