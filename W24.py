import pandas as pd
import numpy as np
import openpyxl 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Lấy dữ liệu từ CMS
mPath = r'E:\Onedrive\9.Packet A1\9.QA-QC-WE\QUERY_W24.xlsm'
df = pd.read_excel(mPath,sheet_name=0)

# Lấy danh sách các số báo cáo cần tạo
Path_report_list = r'E:\Onedrive\9.Packet A1\9.QA-QC-WE\QUERY_W24.xlsm'
df_report = pd.read_excel(Path_report_list, sheet_name=3)
report_dic = df_report.to_dict()
report_list = list(report_dic['list_report'].values())

# Tao hàm tạo báo cáo
def tao_bao_cao(report_1):
    # Làm thử 1 báo cáo
    # report_1 = report_list[0]
    df_rp1 = df.loc[df['W24']==report_1]
    rows = dataframe_to_rows(df_rp1,index=False, header=False)
    # Load from vào openpyxl
    m_path_form = r'E:\\Onedrive\\9.Packet A1\\9.QA-QC-WE\\2.Visual\\Welding Form W24.xlsx'
    wb = load_workbook(m_path_form, read_only= False)
    ws = wb.worksheets[0]
    # Ẩn dòng thừa
    for h in range(10+len(df_rp1.index),209):
        ws.row_dimensions[h].hidden = True
    i = 9
    for row in rows:
        for y in range(1,18):
            ws.cell(column=y,row=i, value=row[y-1])
        i = i+1

    ws.cell(column=17, row= 4, value= row[17])
    # ws.cell(column=2, row= 5, value= row[11])
    
    # Thêm footer cho trang in
    ws.oddFooter.center.text = "Page &[Page] of &N"
    ws.oddFooter.center.size = 8

    # Luu file
    wb.save(r'E:\Onedrive\9.Packet A1\9.QA-QC-WE\17.W24_report' + '\\\\'+ report_1 +'.xlsx')
    wb.close()

for rp in report_list:
    tao_bao_cao(rp)